import hashlib
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.ext.asyncio import AsyncSession

from src.errors import ApiError
from src.repositories.excel_stock_repository import (
    ActiveSnapshot,
    ExcelStockRepository,
    ParsedStockRow,
)

# tools_spec.md "Excel source of truth" column mapping.
REQUIRED_COLUMNS = ("sku", "stock", "unit", "price")
VALID_UNITS = {"pieces", "meters", "kg", "liters", "boxes"}
MAX_ROWS = 500


def _cell_str(value: object) -> str:
    return "" if value is None else str(value).strip()


def _parse_int(value: object, field: str, row_num: int) -> int:
    raw = _cell_str(value)
    try:
        as_float = float(raw)
    except ValueError as exc:
        message = f"Row {row_num}: {field} must be a number"
        raise ApiError(400, "BAD_REQUEST", message, field=field) from exc
    if as_float != int(as_float):
        message = f"Row {row_num}: {field} must be a whole number"
        raise ApiError(400, "BAD_REQUEST", message, field=field)
    parsed = int(as_float)
    if parsed < 0:
        message = f"Row {row_num}: {field} cannot be negative"
        raise ApiError(400, "BAD_REQUEST", message, field=field)
    return parsed


def _parse_price(value: object, row_num: int) -> Decimal:
    try:
        price = Decimal(_cell_str(value))
    except InvalidOperation as exc:
        message = f"Row {row_num}: Price must be a number"
        raise ApiError(400, "BAD_REQUEST", message, field="price") from exc
    if price < 0:
        message = f"Row {row_num}: Price cannot be negative"
        raise ApiError(400, "BAD_REQUEST", message, field="price")
    return price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _parse_unit(value: object, row_num: int) -> str:
    unit = _cell_str(value).lower()
    if unit not in VALID_UNITS:
        allowed = ", ".join(sorted(VALID_UNITS))
        message = f"Row {row_num}: Unit must be one of {allowed}"
        raise ApiError(400, "BAD_REQUEST", message, field="unit")
    return unit


def _column_index(header: list[str], name: str) -> int | None:
    try:
        return header.index(name)
    except ValueError:
        return None


def _parse_rows(file_bytes: bytes) -> list[ParsedStockRow]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise ApiError(400, "BAD_REQUEST", "File is not a valid .xlsx workbook") from exc

    sheet = workbook.active
    if sheet is None:
        raise ApiError(400, "BAD_REQUEST", "Workbook has no sheets")

    all_rows = list(sheet.iter_rows(values_only=True))
    if len(all_rows) == 0:
        raise ApiError(400, "BAD_REQUEST", "File is empty")

    header = [_cell_str(cell).lower() for cell in all_rows[0]]
    for column in REQUIRED_COLUMNS:
        if _column_index(header, column) is None:
            message = f"Missing required column: {column}"
            raise ApiError(400, "BAD_REQUEST", message, field="header")

    sku_idx = _column_index(header, "sku")
    stock_idx = _column_index(header, "stock")
    unit_idx = _column_index(header, "unit")
    price_idx = _column_index(header, "price")
    aliases_idx = _column_index(header, "aliases")
    reorder_idx = _column_index(header, "reorder threshold")
    assert sku_idx is not None
    assert stock_idx is not None
    assert unit_idx is not None
    assert price_idx is not None

    parsed_rows: list[ParsedStockRow] = []
    for offset, row in enumerate(all_rows[1:]):
        row_num = offset + 2  # 1-based, header is row 1
        if all(_cell_str(cell) == "" for cell in row):
            continue

        sku_canonical = _cell_str(row[sku_idx]) if sku_idx < len(row) else ""
        if sku_canonical == "":
            raise ApiError(400, "BAD_REQUEST", f"Row {row_num}: SKU is required", field="sku")

        has_aliases_col = aliases_idx is not None and aliases_idx < len(row)
        aliases_raw = (
            _cell_str(row[aliases_idx]) if has_aliases_col and aliases_idx is not None else ""
        )
        aliases = [a.strip() for a in aliases_raw.split(",") if a.strip() != ""]

        has_reorder_col = reorder_idx is not None and reorder_idx < len(row)
        reorder_raw = row[reorder_idx] if has_reorder_col and reorder_idx is not None else None
        reorder_threshold = (
            _parse_int(reorder_raw, "Reorder Threshold", row_num)
            if _cell_str(reorder_raw) != ""
            else 0
        )

        parsed_rows.append(
            ParsedStockRow(
                sku_canonical=sku_canonical,
                sku_aliases=aliases,
                stock=_parse_int(row[stock_idx], "Stock", row_num),
                unit=_parse_unit(row[unit_idx], row_num),
                price_per_unit=_parse_price(row[price_idx], row_num),
                reorder_threshold=reorder_threshold,
            )
        )

    if len(parsed_rows) == 0:
        raise ApiError(400, "BAD_REQUEST", "File has no data rows")
    if len(parsed_rows) > MAX_ROWS:
        raise ApiError(400, "BAD_REQUEST", f"Too many rows (max {MAX_ROWS})")

    return parsed_rows


async def ingest(
    db: AsyncSession, sme_id: str, file_bytes: bytes, filename: str
) -> tuple[ActiveSnapshot, bool]:
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    repo = ExcelStockRepository(db)

    active = await repo.get_active_snapshot(sme_id)
    if active is not None and active.snapshot_hash == file_hash:
        return active, True

    rows = _parse_rows(file_bytes)
    return await repo.replace_snapshot(sme_id, file_hash, filename, rows), False
